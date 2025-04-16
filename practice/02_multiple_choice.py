from helper_functions import clear_screen
clear_screen()

# ===============
# MULTIPLE CHOICE
# ===============

'''
OVERVIEW
--------

This file has a bunch of functions in it. You can expect to see these functions
on the multiple choice portion.
'''

# ======
# PANDAS
# ======

'''
Given a dataframe, you should know what .iloc does and .loc
'''

import pandas as pd

nba_teams = {
    "Team": ["Boston Celtics", "Denver Nuggets", "Golden State Warriors", "Miami Heat", "Milwaukee Bucks"],
    "City": ["Boston", "Denver", "San Francisco", "Miami", "Milwaukee"],
    "Wins": [64, 57, 46, 44, 49],
    "Losses": [18, 25, 36, 38, 33]
}

df = pd.DataFrame(nba_teams)
print(df)

# what will this print?
print()
print(df.iloc[3,1])

# what will this print?
print()
print(df.loc[2, "Wins"])

# what could you write to get the team from Denver without knowing
# which row that is?
print()
print(df.loc[df["City"]=="Denver"])


# ========
# OPENPYXL
# ========
'''
You should know how active sheets work and how .iter_rows() works.
'''

import openpyxl

wb = openpyxl.load_workbook("mock_grades.xlsx")

for row in wb.active.iter_rows(min_row=2,min_col=2, values_only=True):
    print(row)



# ============
# DICTIONARIES
# ============

'''
How do you get data out of a dictionary? How do you put another key value
pair into a dictionary?
'''
example_dict = {"Bob": 35, "Jane": 40}

# what does this do?
example_dict["Alice"] = 25

# ============
# ADVANCED OOP
# ============

'''
Know how to access objects inside of other objects
Be able to recognize inheritance
'''


# Does this example use inhertance? How can you tell?

class Employee:
    def __init__(self, name):
        self.name = name

class Manager(Employee):
    def __init__(self, name, department):
        super().__init__(name)
        self.department = department


# Does this example use inheritance? How can you tell?

class Dog:
    def __init__(self, dog_name = "Doggy", age = 5):
        self.dog_name = dog_name
        self.age = age

class DogOwner:
    def __init__(self, owner_name):
        self.__owner_name = owner_name
        self.owned_dog = Dog()

do_1 = DogOwner("Bob")

# How could I access Bob's dog's name? Is it even possible?

# Does this work?

try:
    print(do_1.__owner_name)
except:
    pass

# ======================
# RETURNING IN FUNCTIONS
# ======================

# What would print out here?

def combined_names(first_name, last_name):
    full_name = f"{first_name} {last_name}"

print(combined_names("Jimmer", "Fredette"))

# ===========================
# CONTINUE AND BREAK IN LOOPS
# ===========================

# what will print here?

for num in range(5):
    if num == 2:
        continue
    elif num == 4:
        break
    print(num)

# ============
# PEEWEE STUFF
# ============

'''
See the previos file for more in depth peewee practice. For multple choice
you should understand:

- What the .create function does
- Why you would bother overriding the .create function
- How do actually update an object/row in the database and not just in python.

'''